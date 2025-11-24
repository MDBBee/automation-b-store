import openpyxl


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]

    row_count = 0
    for row in sheet.iter_rows(values_only=True):
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            row_count += 1

    return row_count


def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_column)


def readData(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum, column=columnno).value


def writeData(file, sheetName, rownum, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(file)
